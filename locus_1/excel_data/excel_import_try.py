import openpyxl


book = openpyxl.open('real_data.xlsx', read_only=True)
sheet = book.active

# print(sheet.max_row)
# print(sheet.max_column)

# print(sheet[2][sheet.max_column-1])


"""
for row_index in range(3, 24):
    print(sheet[row_index][3].value)
"""
"""
for column_index in range(1, sheet.max_column, 2):
    for row_index in range(3, sheet.max_row + 1):
        print(sheet[row_index][0].value, sheet[row_index][column_index].value, sheet[row_index][column_index + 1].value)
"""
for column_index in range(1, 5, 2):
    for row_index in range(3, sheet.max_row + 1):
        print(sheet[row_index][0].value, sheet[row_index][column_index].value, sheet[row_index][column_index + 1].value)
